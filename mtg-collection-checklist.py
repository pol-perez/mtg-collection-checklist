import requests
import time

from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle


def search_database(type):
    card_list = []
    set_list = []
    url_list = []
    page = 1

    while True:
        # Search Scryfall database using query parameters
        payload = {
            'q': 'type:{} lang:en include:extras unique:prints'.format(type),
            'page': page,
            'order': 'name',
            'as': 'checklist',
        }
        r = requests.get('https://scryfall.com/search', params=payload)

        # Check if page exists
        if r.status_code == 200:
            soup = BeautifulSoup(r.text, 'lxml')
            checklist = soup.find('table', class_='checklist')

            # Find all cards and get card data (name, artist, set, and link)
            for card in checklist.tbody.find_all('tr'):
                # Data variable is a list containing card name, type, and artist
                data = card.find_all('td', class_='ellipsis')
                set_ = card.find('td').a.abbr['title']
                url = card.find('td').a['href']

                card_list.append((data[0].a.text, data[2].a.text))
                set_list.append(set_)
                url_list.append(url)

            # Add delay and go to next page
            print('Page {} parsing completed'.format(page))
            time.sleep(3)
            page += 1
        else:
            break

    return card_list, set_list, url_list


def process_card_data(card_list, set_list, url_list):
    # Use zip to merge the lists, and zip* to unmerge after sorting
    card_list, set_list, url_list = zip(
        *sorted(list(zip(card_list, set_list, url_list))))

    # Get unique art for each card by removing duplicates with same artist
    unique_card_list = list(dict.fromkeys(card_list))

    # Get index of first instance of each unique card
    index_list = []
    for card in unique_card_list:
        index = card_list.index(card)
        index_list.append(index)

    # Get sets of each unique card by slicing the set list
    unique_set_list = []
    unique_url_list = []
    for i in range(len(index_list)):
        # Check if not final iteration
        if i != len(index_list)-1:
            set_ = set_list[index_list[i]:index_list[i+1]]
        else:
            set_ = set_list[index_list[i]:]

        unique_set_list.append(set_)

        # Get one URL for each unique card
        single_url = 'https://scryfall.com' + url_list[index_list[i]]
        unique_url_list.append(single_url)

    # Generate final card list
    output_list = []
    for i in range(len(unique_card_list)):
        data = (
            unique_card_list[i][0], unique_card_list[i][1], unique_set_list[i], unique_url_list[i])
        output_list.append(data)

    return output_list


def generate_checklist(list, type_):
    # Create spreadsheet
    workbook = Workbook()
    sheet = workbook.active

    # Create cell styles
    default = NamedStyle(name='default')
    default.alignment = Alignment(
        horizontal='center', vertical='center', wrap_text=True)
    side = Side(border_style='thin')
    default.border = Border(top=side, right=side, bottom=side, left=side)
    header_font = Font(bold=True)
    url_font = Font(color='000000FF', underline='single')
    pattern = PatternFill(patternType='solid', fgColor='00C0C0C0')

   # Print headers and add styling
    headers = ['NAME', 'ARTIST', 'SET', 'LINK']
    row = 1
    column = 1
    for header in headers:
        sheet.cell(row=row, column=column).value = header
        sheet.cell(row=row, column=column).style = default
        sheet.cell(row=row, column=column).font = header_font
        sheet.cell(row=row, column=column).fill = pattern
        column += 1
    row += 1

    # Print card data
    for r in range(len(list)):
        sheet.cell(row=row, column=1).value = list[r][0]
        sheet.cell(row=row, column=2).value = list[r][1]
        sheet.cell(row=row, column=3).value = " • ".join(list[r][2])
        sheet.cell(row=row, column=4).hyperlink = list[r][3]
        row += 1

    # Add styling
    row = 2
    for r in range(len(list)):
        for c in range(1, len(list[r])+1):
            sheet.cell(row=row, column=c).style = default
            if r % 2 == 1:
                sheet.cell(row=row, column=c).fill = pattern
        sheet.cell(row=row, column=4).font = url_font
        row += 1

    # Set column width
    for column in 'ABCD':
        sheet.column_dimensions[column].width = 30

    # Save spreadsheet file
    title = "{}_Checklist.xlsx".format(type_.capitalize())
    workbook.save(filename=title)


# Input card type
type_ = input("Search for card type: ")

# Execute functions
card_list, set_list, url_list = search_database(type_)
output_list = process_card_data(card_list, set_list, url_list)
generate_checklist(output_list, type_)
